"""
Semantic SEO Tool — Production Backend
======================================
Lead Architect: Production-Grade Flask API
- Nara Router API integration (7M token budget)
- Pydantic v2 schema validation
- Multi-input: URL, .docx, .txt, source code
- Chain-of-Thought prompt engineering
- Zero API key exposure to frontend
"""

import os
import re
import io
import json
import base64
import logging
import traceback
from typing import Optional
from collections import Counter

import requests
from dotenv import load_dotenv
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS

# ── Pydantic v2 ──────────────────────────────────────────────────────────────
from pydantic import BaseModel, Field, field_validator

# ── Document parsers ─────────────────────────────────────────────────────────
from bs4 import BeautifulSoup
import docx
import openpyxl
import csv

# ── Excel export ─────────────────────────────────────────────────────────────
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# 0. Bootstrap
# ─────────────────────────────────────────────────────────────────────────────

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("semantic-seo")

app = Flask(__name__, static_folder=".", static_url_path="")
CORS(app)

NARA_API_KEY    = os.environ.get("NARA_API_KEY", "")
NARA_API_URL    = os.environ.get("NARA_API_URL", "https://api.nara.tools/v1/chat/completions")
NARA_MODEL      = os.environ.get("NARA_MODEL", "claude-sonnet-4-5")

MAX_CHARS       = 25_000   # token budget allows deep extraction
MAX_COMPETITORS = 5        # cap to keep latency manageable

if not NARA_API_KEY:
    logger.warning("NARA_API_KEY is not set — analysis endpoint will return 500.")


# ─────────────────────────────────────────────────────────────────────────────
# 1. Pydantic Schemas  (single source of truth for AI output)
# ─────────────────────────────────────────────────────────────────────────────

class EntityItem(BaseModel):
    """Single semantic entity extracted from content."""
    name:       str = Field(..., description="Entity name in original language")
    entity_type: str = Field(..., alias="type",
                             description="Person | Organization | Place | Concept | Service | Event | Product")
    importance: str = Field(default="medium",
                            description="high | medium | low based on Search Intent relevance")
    salience:   float = Field(default=0.5, ge=0.0, le=1.0,
                              description="Semantic salience score 0–1")
    mentions:   int   = Field(default=1, ge=0)

    @field_validator("entity_type")
    @classmethod
    def normalize_type(cls, raw: str) -> str:
        allowed = {"Person", "Organization", "Place", "Concept",
                   "Service", "Event", "Product"}
        cleaned = raw.strip().title()
        return cleaned if cleaned in allowed else "Concept"

    @field_validator("importance")
    @classmethod
    def normalize_importance(cls, raw: str) -> str:
        raw = raw.strip().lower()
        return raw if raw in {"high", "medium", "low"} else "medium"

    model_config = {"populate_by_name": True}


class SemanticCluster(BaseModel):
    """Group of thematically related entities."""
    cluster_name: str
    entities:     list[EntityItem]


class SourceAnalysis(BaseModel):
    """Full entity analysis for one content source."""
    source_label: str                          # "موقعي" or "المنافس N"
    clusters:     list[SemanticCluster]
    raw_entities: list[EntityItem] = Field(default_factory=list)


class SemanticGap(BaseModel):
    """Entity present in competitors but absent/weak in client site."""
    name:               str
    entity_type:        str = Field(alias="type")
    competitor_count:   int   = Field(description="How many competitors mention it")
    total_mentions:     int   = Field(description="Sum of competitor mentions")
    priority:           str   = Field(description="critical | high | medium")

    model_config = {"populate_by_name": True}


class FullAnalysisResult(BaseModel):
    """Root response model returned to the frontend."""
    master_entities:   list[EntityItem]
    my_site_entities:  list[EntityItem]
    competitor_breakdowns: list[SourceAnalysis]
    semantic_gaps:     list[SemanticGap]
    content_brief:     str = ""
    has_my_article:    bool = False
    has_competitors:   bool = False


# ─────────────────────────────────────────────────────────────────────────────
# 2. Text Extraction Layer
# ─────────────────────────────────────────────────────────────────────────────

def _strip_navigation(soup: BeautifulSoup) -> str:
    """Remove nav/header/footer/aside from HTML before extraction."""
    for tag in soup.find_all(["nav", "header", "footer", "aside",
                               "script", "style", "noscript"]):
        tag.decompose()
    # Also drop elements with common navigation class/id patterns
    nav_patterns = re.compile(
        r"(nav|menu|sidebar|breadcrumb|footer|header|cookie|banner)",
        re.I
    )
    for tag in soup.find_all(True):
        classes = " ".join(tag.get("class", []))
        tag_id  = tag.get("id", "")
        if nav_patterns.search(classes) or nav_patterns.search(tag_id):
            tag.decompose()
    return soup.get_text(separator=" ", strip=True)


def _sanitize_code(raw: str) -> str:
    """Strip code noise, keep identifiers and comments for entity extraction."""
    # Remove string literals
    raw = re.sub(r'"[^"]*"', " ", raw)
    raw = re.sub(r"'[^']*'", " ", raw)
    # Remove URLs
    raw = re.sub(r"https?://\S+", " ", raw)
    # Collapse whitespace
    raw = re.sub(r"\s+", " ", raw).strip()
    return raw


def extract_text_from_url(url: str) -> str:
    """Fetch URL and extract main content via BeautifulSoup."""
    try:
        resp = requests.get(url, timeout=15, headers={
            "User-Agent": "Mozilla/5.0 (SEO-Tool/1.0)"
        })
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        # Prefer <main> or <article> if available
        main = soup.find("main") or soup.find("article")
        if main:
            return _strip_navigation(main)[:MAX_CHARS]
        return _strip_navigation(soup)[:MAX_CHARS]
    except Exception as exc:
        logger.error("URL fetch failed for %s: %s", url, exc)
        raise ValueError(f"تعذّر جلب الرابط: {url}") from exc


def extract_text_from_file(file_obj) -> str:
    """Dispatch to the right extractor based on file extension."""
    filename = getattr(file_obj, "filename", "unknown").lower()

    try:
        if filename.endswith(".txt"):
            raw = file_obj.read().decode("utf-8", errors="ignore")
            return raw[:MAX_CHARS]

        if filename.endswith(".docx"):
            doc   = docx.Document(file_obj)
            lines = [p.text for p in doc.paragraphs if p.text.strip()]
            return " ".join(lines)[:MAX_CHARS]

        if filename.endswith(".csv"):
            content = file_obj.read().decode("utf-8", errors="ignore")
            reader  = csv.reader(io.StringIO(content))
            rows    = [" ".join(cell for cell in row) for row in reader]
            return " ".join(rows)[:MAX_CHARS]

        if filename.endswith((".xlsx", ".xls")):
            wb   = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
            rows = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    rows.append(" ".join(str(c) for c in row if c is not None))
            return " ".join(rows)[:MAX_CHARS]

        if filename.endswith((".py", ".js", ".ts", ".html", ".css",
                               ".jsx", ".tsx", ".java", ".cpp", ".cs")):
            raw = file_obj.read().decode("utf-8", errors="ignore")
            return _sanitize_code(raw)[:MAX_CHARS]

        # Fallback — treat as plain text
        raw = file_obj.read().decode("utf-8", errors="ignore")
        return raw[:MAX_CHARS]

    except Exception as exc:
        logger.error("File extraction failed for %s: %s", filename, exc)
        raise ValueError(f"تعذّر قراءة الملف: {filename}") from exc


# ─────────────────────────────────────────────────────────────────────────────
# 3. Nara API — Chain-of-Thought Semantic Extraction
# ─────────────────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """أنت Lead Semantic SEO Architect من الدرجة الأولى.
مهمتك: تحليل النصوص المقدمة واستخراج الكيانات الدلالية كخبير يرى الأفكار كـ "علاقات بين كيانات" (Entities & Relationships)، وليس مجرد كلمات منفصلة.

قواعد التحليل (اتبعها بدقة):
1. تجاهل تماماً: عناصر التنقل (Navigation)، الفوتر (Footer)، الهيدر (Header)، الروابط الجانبية، وأي نص ليس من Main Content.
2. ركّز على: الكيانات ذات الصلة بـ Search Intent والموضوع المحوري للنص.
3. صنّف الأهمية (importance) بناءً على: قرب الكيان من الموضوع الرئيسي، وتكراره، وعلاقته بنية البحث.
4. أعط كل كيان درجة salience (0.0–1.0) تعكس ثقله الدلالي النسبي في النص.
5. جمّع الكيانات في عناقيد دلالية (Topic Clusters) متجانسة.

أنواع الكيانات المسموح بها فقط:
Person | Organization | Place | Concept | Service | Event | Product

الإخراج: JSON خالص بدون أي نص إضافي، يطابق الهيكل المطلوب بدقة."""


def _build_extraction_prompt(label: str, text: str) -> str:
    """Build the per-source extraction prompt with CoT structure."""
    return f"""## المصدر: {label}

### النص المراد تحليله:
{text}

### المطلوب:
فكّر خطوة بخطوة (Chain of Thought):
1. ما الموضوع الرئيسي لهذا النص؟
2. ما الكيانات المحورية (high importance) التي يدور حولها؟
3. ما الكيانات الداعمة (medium/low importance)؟
4. كيف تتجمع هذه الكيانات في عناقيد دلالية؟

ثم أعِد JSON يطابق هذا الهيكل تماماً:
{{
  "source_label": "{label}",
  "clusters": [
    {{
      "cluster_name": "اسم العنقود",
      "entities": [
        {{
          "name": "اسم الكيان",
          "type": "Concept",
          "importance": "high",
          "salience": 0.85,
          "mentions": 3
        }}
      ]
    }}
  ]
}}"""


def call_nara_api(prompt: str, max_tokens: int = 4096) -> str:
    """
    Call Nara Router API using OpenAI-compatible /chat/completions endpoint.
    Returns raw text content from the response.
    """
    if not NARA_API_KEY:
        raise RuntimeError("NARA_API_KEY غير مضبوط في البيئة.")

    payload = {
        "model":       NARA_MODEL,
        "max_tokens":  max_tokens,
        "temperature": 0.1,
        "messages": [
            {"role": "system",  "content": SYSTEM_PROMPT},
            {"role": "user",    "content": prompt},
        ],
    }

    headers = {
        "Authorization": f"Bearer {NARA_API_KEY}",
        "Content-Type":  "application/json",
    }

    try:
        resp = requests.post(
            NARA_API_URL,
            headers=headers,
            json=payload,
            timeout=120,
        )
        resp.raise_for_status()
    except requests.exceptions.Timeout:
        raise RuntimeError("انتهت مهلة الاتصال بـ Nara API (120 ثانية).")
    except requests.exceptions.HTTPError as exc:
        status = exc.response.status_code if exc.response is not None else "?"
        body   = exc.response.text[:300] if exc.response is not None else ""
        raise RuntimeError(f"Nara API HTTP {status}: {body}") from exc
    except requests.exceptions.RequestException as exc:
        raise RuntimeError(f"خطأ في الاتصال بـ Nara API: {exc}") from exc

    data = resp.json()
    # Standard OpenAI response shape
    choices = data.get("choices") or []
    if not choices:
        raise RuntimeError(f"Nara API لم يُعِد أي choices. الرد: {str(data)[:300]}")

    content = choices[0].get("message", {}).get("content", "")
    if not content:
        raise RuntimeError("Nara API أعاد محتوى فارغاً.")
    return content


def _parse_json_from_response(raw: str) -> dict:
    """Extract and parse JSON from a response that may include surrounding text."""
    # Try direct parse first
    try:
        return json.loads(raw.strip())
    except json.JSONDecodeError:
        pass
    # Fallback: extract first {...} block
    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            pass
    raise ValueError(f"لم يمكن استخراج JSON من الرد: {raw[:200]}")


def extract_entities_for_source(label: str, text: str) -> SourceAnalysis:
    """Run deep semantic extraction for a single source."""
    prompt  = _build_extraction_prompt(label, text)
    raw_out = call_nara_api(prompt, max_tokens=4096)
    data    = _parse_json_from_response(raw_out)

    # Validate & coerce via Pydantic
    analysis = SourceAnalysis.model_validate(data)

    # Flatten entities for easy access
    flat: list[EntityItem] = []
    for cluster in analysis.clusters:
        for ent in cluster.entities:
            if len(ent.name.strip()) > 2:
                flat.append(ent)
    analysis.raw_entities = flat
    return analysis


# ─────────────────────────────────────────────────────────────────────────────
# 4. Semantic Gap Analysis
# ─────────────────────────────────────────────────────────────────────────────

def compute_semantic_gaps(
    my_entities:   list[EntityItem],
    comp_analyses: list[SourceAnalysis],
) -> list[SemanticGap]:
    """Identify high-value entities in competitors missing from client site."""

    my_names: set[str] = {e.name.strip().lower() for e in my_entities}

    # Aggregate competitor mentions per entity name
    comp_counter:  Counter = Counter()
    comp_sources:  Counter = Counter()
    comp_type_map: dict[str, str] = {}

    for comp_analysis in comp_analyses:
        seen_in_this = set()
        for ent in comp_analysis.raw_entities:
            key = ent.name.strip().lower()
            comp_counter[key] += ent.mentions
            comp_type_map[key] = comp_type_map.get(key, ent.entity_type)
            if key not in seen_in_this:
                comp_sources[key] += 1
                seen_in_this.add(key)

    gaps: list[SemanticGap] = []
    for entity_name_lower, total_mentions in comp_counter.most_common():
        if entity_name_lower in my_names:
            continue  # already covered by client

        # Restore original casing (use first seen)
        original_name = entity_name_lower  # best effort
        for comp in comp_analyses:
            for ent in comp.raw_entities:
                if ent.name.strip().lower() == entity_name_lower:
                    original_name = ent.name
                    break

        src_count = comp_sources[entity_name_lower]
        n_comps   = len(comp_analyses)

        # Priority heuristic
        coverage_ratio = src_count / max(n_comps, 1)
        if coverage_ratio >= 0.7 or total_mentions >= 8:
            priority = "critical"
        elif coverage_ratio >= 0.4 or total_mentions >= 4:
            priority = "high"
        else:
            priority = "medium"

        gaps.append(SemanticGap(
            name=original_name,
            type=comp_type_map.get(entity_name_lower, "Concept"),
            competitor_count=src_count,
            total_mentions=total_mentions,
            priority=priority,
        ))

    # Sort: critical → high → medium, then by total_mentions
    priority_order = {"critical": 0, "high": 1, "medium": 2}
    gaps.sort(key=lambda g: (priority_order[g.priority], -g.total_mentions))
    return gaps


# ─────────────────────────────────────────────────────────────────────────────
# 5. Content Brief Generation
# ─────────────────────────────────────────────────────────────────────────────

def generate_content_brief(gaps: list[SemanticGap], topic_context: str = "") -> str:
    """Generate an SEO content brief targeting the top semantic gaps."""
    if not gaps:
        return ""

    critical = [g.name for g in gaps if g.priority == "critical"][:8]
    high     = [g.name for g in gaps if g.priority == "high"][:7]

    prompt = f"""أنت كاتب محتوى سيو متخصص.

الموضوع العام: {topic_context or 'المحتوى المحلل'}

الفجوات الدلالية الحرجة التي يجب سدّها:
{', '.join(critical) if critical else 'لا يوجد'}

الفجوات عالية الأولوية:
{', '.join(high) if high else 'لا يوجد'}

المطلوب: اكتب هيكل مقال سيو متكاملاً بـ H2 وH3 وH4 يسدّ هذه الفجوات.
- ابدأ كل قسم بكلمة مفتاحية طويلة الذيل (Long-tail keyword) مناسبة.
- أشر إلى فرص الـ Internal Linking بين الأقسام.
- قدّر عدد الكلمات الموصى به لكل قسم.
- أخرج Markdown فقط بدون مقدمات."""

    try:
        return call_nara_api(prompt, max_tokens=2048)
    except Exception as exc:
        logger.warning("Content brief generation failed: %s", exc)
        return ""


# ─────────────────────────────────────────────────────────────────────────────
# 6. Excel Export
# ─────────────────────────────────────────────────────────────────────────────

_BLUE_FILL   = PatternFill("solid", fgColor="0101B1")
_ROSE_FILL   = PatternFill("solid", fgColor="FF4B6E")
_GREEN_FILL  = PatternFill("solid", fgColor="059669")
_GRAY_FILL   = PatternFill("solid", fgColor="F1F5F9")
_WHITE_FONT  = Font(color="FFFFFF", bold=True)
_DARK_FONT   = Font(color="1E293B")
_BOLD_DARK   = Font(color="1E293B", bold=True)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RIGHT_ALIGN  = Alignment(horizontal="right", vertical="center", wrap_text=True)

def _header_row(ws, headers: list[str], fill: PatternFill, row: int = 1):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill        = fill
        cell.font        = _WHITE_FONT
        cell.alignment   = _CENTER_ALIGN


def build_excel_report(result: FullAnalysisResult) -> bytes:
    """Build a styled multi-sheet Excel report."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Sheet 1: Master Entities ─────────────────────────────────────────────
    ws1 = wb.create_sheet("الكيانات الشاملة")
    ws1.right_to_left = True
    _header_row(ws1, ["الكيان", "النوع", "الأهمية", "درجة الدلالة", "في موقعي", "في المنافسين"], _BLUE_FILL)
    my_names = {e.name.strip().lower() for e in result.my_site_entities}
    comp_names: Counter = Counter()
    for comp in result.competitor_breakdowns:
        for ent in comp.raw_entities:
            comp_names[ent.name.strip().lower()] += ent.mentions

    for row_i, ent in enumerate(result.master_entities, start=2):
        key    = ent.name.strip().lower()
        in_my  = "✓" if key in my_names else "—"
        in_comp = comp_names.get(key, 0)
        row_fill = _GRAY_FILL if row_i % 2 == 0 else None
        for col_i, val in enumerate([ent.name, ent.entity_type, ent.importance,
                                      f"{ent.salience:.2f}", in_my, in_comp], start=1):
            cell = ws1.cell(row=row_i, column=col_i, value=val)
            cell.alignment = _RIGHT_ALIGN
            if row_fill:
                cell.fill = row_fill
    ws1.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E", "F"]:
        ws1.column_dimensions[col].width = 16

    # ── Sheet 2: Semantic Gaps ────────────────────────────────────────────────
    ws2 = wb.create_sheet("الفجوات الدلالية")
    ws2.right_to_left = True
    _header_row(ws2, ["الكيان", "النوع", "الأولوية", "عدد المنافسين", "إجمالي التكرار"], _ROSE_FILL)
    priority_colors = {"critical": "FEE2E2", "high": "FEF3C7", "medium": "F0FDF4"}
    for row_i, gap in enumerate(result.semantic_gaps, start=2):
        row_fill = PatternFill("solid", fgColor=priority_colors.get(gap.priority, "FFFFFF"))
        for col_i, val in enumerate([gap.name, gap.entity_type, gap.priority,
                                      gap.competitor_count, gap.total_mentions], start=1):
            cell = ws2.cell(row=row_i, column=col_i, value=val)
            cell.alignment = _RIGHT_ALIGN
            cell.fill = row_fill
    ws2.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E"]:
        ws2.column_dimensions[col].width = 18

    # ── Sheet 3: Content Brief ────────────────────────────────────────────────
    if result.content_brief:
        ws3 = wb.create_sheet("خطة المحتوى")
        ws3.right_to_left = True
        _header_row(ws3, ["هيكل المقال المقترح"], _GREEN_FILL)
        ws3.cell(row=2, column=1, value=result.content_brief).alignment = Alignment(wrap_text=True)
        ws3.column_dimensions["A"].width = 80
        ws3.row_dimensions[2].height = 400

    # ── Sheet 4: Competitor Breakdown ─────────────────────────────────────────
    for comp in result.competitor_breakdowns:
        safe_name = re.sub(r"[\\/*?\[\]:]", "", comp.source_label)[:28]
        ws_c = wb.create_sheet(safe_name)
        ws_c.right_to_left = True
        _header_row(ws_c, ["العنقود الدلالي", "الكيان", "النوع", "الأهمية", "التكرار"], _BLUE_FILL)
        row_i = 2
        for cluster in comp.clusters:
            for ent in cluster.entities:
                for col_i, val in enumerate(
                    [cluster.cluster_name, ent.name, ent.entity_type,
                     ent.importance, ent.mentions], start=1
                ):
                    cell = ws_c.cell(row=row_i, column=col_i, value=val)
                    cell.alignment = _RIGHT_ALIGN
                    if row_i % 2 == 0:
                        cell.fill = _GRAY_FILL
                row_i += 1
        ws_c.column_dimensions["A"].width = 25
        ws_c.column_dimensions["B"].width = 30
        for col in ["C", "D", "E"]:
            ws_c.column_dimensions[col].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# 7. Flask Routes
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/")
def serve_index():
    return send_from_directory(".", "index.html")


@app.route("/health")
def health():
    return jsonify({
        "status":    "ok",
        "api_ready": bool(NARA_API_KEY),
        "model":     NARA_MODEL,
    })


@app.route("/analyze", methods=["POST"])
def analyze():
    """
    Main analysis endpoint.

    Accepts multipart/form-data with:
        - my_article: file (.txt, .docx, .csv, .xlsx, source code)
        - my_url:     URL string (alternative to file)
        - competitors: one or more files
        - comp_urls:  one or more URL strings
    """
    if not NARA_API_KEY:
        return jsonify({"error": "NARA_API_KEY غير مضبوط على السيرفر. تواصل مع المشرف."}), 500

    # ── Gather inputs ─────────────────────────────────────────────────────────
    my_file     = request.files.get("my_article")
    my_url      = request.form.get("my_url", "").strip()
    comp_files  = [f for f in request.files.getlist("competitors") if f.filename][:MAX_COMPETITORS]
    comp_urls   = [u.strip() for u in request.form.getlist("comp_urls") if u.strip()][:MAX_COMPETITORS]

    has_my_article  = bool((my_file and my_file.filename) or my_url)
    has_competitors = bool(comp_files or comp_urls)

    if not has_my_article and not has_competitors:
        return jsonify({"error": "يرجى رفع ملف أو إدخال رابط على الأقل."}), 400

    # ── Extract texts ─────────────────────────────────────────────────────────
    my_text:    str = ""
    comp_texts: list[tuple[str, str]] = []   # (label, text)

    try:
        if my_url:
            my_text = extract_text_from_url(my_url)
        elif my_file and my_file.filename:
            my_text = extract_text_from_file(my_file)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 422

    for idx, cf in enumerate(comp_files, start=1):
        try:
            comp_texts.append((f"المنافس {idx}", extract_text_from_file(cf)))
        except ValueError as exc:
            logger.warning("Skipping competitor file %s: %s", cf.filename, exc)

    for url in comp_urls:
        idx = len(comp_texts) + 1
        try:
            comp_texts.append((f"المنافس {idx}", extract_text_from_url(url)))
        except ValueError as exc:
            logger.warning("Skipping competitor URL %s: %s", url, exc)

    # ── Run semantic extraction (sequential — Nara API is the bottleneck) ─────
    my_analysis:    Optional[SourceAnalysis] = None
    comp_analyses:  list[SourceAnalysis]     = []

    try:
        if my_text:
            my_analysis = extract_entities_for_source("موقعي", my_text)
    except Exception as exc:
        logger.error("My-site extraction failed: %s\n%s", exc, traceback.format_exc())
        return jsonify({"error": f"فشل تحليل موقعك: {exc}"}), 500

    for label, text in comp_texts:
        try:
            comp_analyses.append(extract_entities_for_source(label, text))
        except Exception as exc:
            logger.warning("Competitor extraction failed for %s: %s", label, exc)

    # ── Build master entity list ──────────────────────────────────────────────
    all_entities: list[EntityItem] = []
    if my_analysis:
        all_entities.extend(my_analysis.raw_entities)
    for comp in comp_analyses:
        all_entities.extend(comp.raw_entities)

    # Deduplicate by name (case-insensitive), keep highest salience
    master_map: dict[str, EntityItem] = {}
    for ent in all_entities:
        key = ent.name.strip().lower()
        if key not in master_map or ent.salience > master_map[key].salience:
            master_map[key] = ent
    master_entities = sorted(master_map.values(), key=lambda e: (-e.salience, e.name))

    # ── Compute gaps ──────────────────────────────────────────────────────────
    my_site_entities = my_analysis.raw_entities if my_analysis else []
    gaps = compute_semantic_gaps(my_site_entities, comp_analyses) if comp_analyses else []

    # ── Content brief ─────────────────────────────────────────────────────────
    topic_context = (my_analysis.clusters[0].cluster_name
                     if my_analysis and my_analysis.clusters else "")
    brief = generate_content_brief(gaps, topic_context) if gaps else ""

    # ── Build result ──────────────────────────────────────────────────────────
    result = FullAnalysisResult(
        master_entities=master_entities,
        my_site_entities=my_site_entities,
        competitor_breakdowns=comp_analyses,
        semantic_gaps=gaps,
        content_brief=brief,
        has_my_article=bool(my_analysis),
        has_competitors=bool(comp_analyses),
    )

    # ── Excel export ──────────────────────────────────────────────────────────
    try:
        excel_bytes = build_excel_report(result)
        excel_b64   = base64.b64encode(excel_bytes).decode("utf-8")
    except Exception as exc:
        logger.warning("Excel build failed: %s", exc)
        excel_b64 = ""

    # Serialize using Pydantic (handles alias + nested models)
    payload = result.model_dump(by_alias=True)
    payload["excel_file"] = excel_b64
    return jsonify(payload)


# ─────────────────────────────────────────────────────────────────────────────
# 8. Entry Point
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port  = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "false").lower() == "true"
    logger.info("Starting Semantic SEO Tool on port %d (debug=%s)", port, debug)
    app.run(host="0.0.0.0", port=port, debug=debug)
